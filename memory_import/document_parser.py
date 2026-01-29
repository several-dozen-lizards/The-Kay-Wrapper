"""
Document Parser for Kay Zero Memory Import System
Handles: .txt, .doc, .docx, .pdf, .json, .xlsx
Extracts text and metadata from various file formats
"""

import os
import re
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from pathlib import Path


class DocumentChunk:
    """Represents a chunk of a document with metadata."""

    def __init__(self, text: str, metadata: Dict, chunk_index: int = 0, total_chunks: int = 1):
        self.text = text
        self.metadata = metadata
        self.chunk_index = chunk_index
        self.total_chunks = total_chunks

    def to_dict(self) -> Dict:
        return {
            "text": self.text,
            "metadata": self.metadata,
            "chunk_index": self.chunk_index,
            "total_chunks": self.total_chunks
        }


class DocumentParser:
    """
    Parses various document formats and extracts text + metadata.
    Supports chunking for large documents.
    """

    def __init__(self, chunk_size: int = 3000, overlap: int = 500):
        """
        Args:
            chunk_size: Maximum characters per chunk
            overlap: Character overlap between chunks for context preservation
        """
        self.chunk_size = chunk_size
        self.overlap = overlap

    def parse_file(self, file_path: str) -> List[DocumentChunk]:
        """
        Parse a file and return chunks with metadata.

        Args:
            file_path: Path to document file

        Returns:
            List of DocumentChunk objects
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Extract metadata
        metadata = self._extract_metadata(path)

        # Extract text based on file type
        extension = path.suffix.lower()

        if extension == '.txt':
            text = self._parse_txt(path)
        elif extension == '.json':
            text = self._parse_json(path)
        elif extension == '.pdf':
            text = self._parse_pdf(path)
        elif extension in ['.doc', '.docx']:
            text = self._parse_docx(path)
        elif extension == '.xlsx':
            text = self._parse_xlsx(path)
        else:
            raise ValueError(f"Unsupported file type: {extension}")

        # Chunk the text
        chunks = self._chunk_text(text, metadata)

        return chunks

    def _extract_metadata(self, path: Path) -> Dict:
        """Extract metadata from file path and properties."""
        metadata = {
            "filename": path.name,
            "file_path": str(path.absolute()),
            "file_type": path.suffix.lower(),
            "file_size": path.stat().st_size,
            "modified_date": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        }

        # Try to extract date from filename
        date = self._extract_date_from_filename(path.name)
        if date:
            metadata["extracted_date"] = date

        return metadata

    def _extract_date_from_filename(self, filename: str) -> Optional[str]:
        """
        Extract date from filename patterns like:
        - transcript_20241023.txt
        - notes_2024-10-23.txt
        - 2024_10_23_journal.txt
        """
        # Pattern 1: YYYYMMDD
        match = re.search(r'(\d{8})', filename)
        if match:
            date_str = match.group(1)
            try:
                date = datetime.strptime(date_str, '%Y%m%d')
                return date.strftime('%Y-%m-%d')
            except ValueError:
                pass

        # Pattern 2: YYYY-MM-DD or YYYY_MM_DD
        match = re.search(r'(\d{4})[-_](\d{2})[-_](\d{2})', filename)
        if match:
            try:
                date = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
                return date.strftime('%Y-%m-%d')
            except ValueError:
                pass

        return None

    def _chunk_text(self, text: str, metadata: Dict) -> List[DocumentChunk]:
        """
        Split text into chunks with overlap.

        Args:
            text: Full document text
            metadata: Document metadata

        Returns:
            List of DocumentChunk objects
        """
        if len(text) <= self.chunk_size:
            return [DocumentChunk(text, metadata, 0, 1)]

        chunks = []
        start = 0
        chunk_index = 0

        while start < len(text):
            end = start + self.chunk_size

            # Try to break at sentence boundary
            if end < len(text):
                # Look for sentence ending within overlap window
                sentence_end = text.rfind('.', end - self.overlap, end)
                if sentence_end > start:
                    end = sentence_end + 1

            chunk_text = text[start:end].strip()

            # Skip empty chunks
            if chunk_text:
                chunk = DocumentChunk(
                    text=chunk_text,
                    metadata=metadata.copy(),
                    chunk_index=chunk_index,
                    total_chunks=-1  # Will update after
                )
                chunks.append(chunk)
                chunk_index += 1

            # Move start position (with overlap)
            start = end - self.overlap if end < len(text) else end

        # Update total_chunks for all chunks
        total = len(chunks)
        for chunk in chunks:
            chunk.total_chunks = total

        return chunks

    def _parse_txt(self, path: Path) -> str:
        """Parse plain text file."""
        encodings = ['utf-8', 'latin-1', 'cp1252']

        for encoding in encodings:
            try:
                with open(path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue

        raise ValueError(f"Could not decode {path} with any standard encoding")

    def _parse_json(self, path: Path) -> str:
        """Parse JSON file and return formatted text."""
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Convert JSON to formatted text
        return json.dumps(data, indent=2)

    def _parse_pdf(self, path: Path) -> str:
        """Parse PDF file using pdfplumber."""
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber not installed. Run: pip install pdfplumber")

        text_parts = []

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)

        return "\n\n".join(text_parts)

    def _parse_docx(self, path: Path) -> str:
        """Parse Word document using python-docx."""
        try:
            from docx import Document
        except ImportError:
            raise ImportError("python-docx not installed. Run: pip install python-docx")

        doc = Document(path)
        text_parts = []

        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)

        return "\n\n".join(text_parts)

    def _parse_xlsx(self, path: Path) -> str:
        """Parse Excel file using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        workbook = load_workbook(path, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===\n")

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)

        return "\n".join(text_parts)

    def parse_directory(self, directory: str, file_extensions: Optional[List[str]] = None) -> List[DocumentChunk]:
        """
        Parse all supported files in a directory.

        Args:
            directory: Directory path
            file_extensions: Optional list of extensions to filter (e.g., ['.txt', '.pdf'])

        Returns:
            List of all DocumentChunk objects from all files
        """
        if file_extensions is None:
            file_extensions = ['.txt', '.json', '.pdf', '.doc', '.docx', '.xlsx']

        all_chunks = []
        directory_path = Path(directory)

        if not directory_path.exists():
            raise FileNotFoundError(f"Directory not found: {directory}")

        # Find all matching files
        for file_path in directory_path.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in file_extensions:
                try:
                    chunks = self.parse_file(str(file_path))
                    all_chunks.extend(chunks)
                    print(f"[PARSER] Parsed {file_path.name}: {len(chunks)} chunks")
                except Exception as e:
                    print(f"[PARSER ERROR] Failed to parse {file_path.name}: {e}")
                    continue

        return all_chunks


# Testing
if __name__ == "__main__":
    parser = DocumentParser()

    # Test with a simple text file
    test_text = "This is a test document.\n\nIt has multiple paragraphs.\n\nAnd should be parsed correctly."

    # Create test file
    test_path = Path("test_document.txt")
    test_path.write_text(test_text)

    try:
        chunks = parser.parse_file(str(test_path))
        print(f"Parsed {len(chunks)} chunk(s)")
        for i, chunk in enumerate(chunks):
            print(f"\nChunk {i+1}/{chunk.total_chunks}:")
            print(f"Text: {chunk.text[:100]}...")
            print(f"Metadata: {chunk.metadata}")
    finally:
        test_path.unlink()  # Clean up
